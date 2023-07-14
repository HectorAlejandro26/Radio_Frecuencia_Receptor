import serial.tools.list_ports as puertos
from serial import Serial, SerialException
from constantes import *
from time import sleep, strftime
from datetime import datetime, timedelta
from os import (
    mkdir,
    path,
    getcwd,
    listdir,
    remove
)
from collections import namedtuple
from typing import Literal, Tuple, Union
from openpyxl import load_workbook, Workbook
from shutil import copy as sh_copy

# NOTE: Solo para debug
from debuging import dbg


class ExcelManager:
    def __init__(self) -> None:
        pass

    def save(self, datos: str) -> None:
        self.mkdirs()

        datos = datos.split('\n')[:-4]

        try:
            libro_trabajo = load_workbook(self.__obtener_ruta_archivo().temp)
            hoja = libro_trabajo.active

        except FileNotFoundError:
            libro_trabajo = Workbook()
            hoja = libro_trabajo.active

            hoja.append(secciones_encabezado)

        datos.extend(
            [self.__tiempo().hora, self.__tiempo().fecha.replace('_', '/')])
        hoja.append(datos)

        libro_trabajo.save(self.__obtener_ruta_archivo().temp)

    def copiar_hacia(self):
        self.mkdirs()
        try:
            sh_copy(self.__obtener_ruta_archivo().temp,
                    self.__obtener_ruta_archivo().registro)
        except PermissionError:
            print('error')
            return False
        else:
            return True

    @staticmethod
    def mkdirs():
        mkdir(ruta_carpeta_registro) if not path.exists(
            ruta_carpeta_registro) else None
        mkdir(ruta_carpeta_temp) if not path.exists(
            ruta_carpeta_temp) else None

    @staticmethod
    def __tiempo(select: Literal['fecha', 'hora'] = None) -> Union[str, Tuple[str, str]]:
        fecha = strftime('%d_%m_%Y')
        hora = strftime('%I:%M:%S') + strftime(' - %p').lower()
        if select == 'fecha':
            return fecha
        elif select == 'hora':
            return hora
        else:
            Tiempo = namedtuple('Tiempo', ['fecha', 'hora'])
            tiempo = Tiempo(fecha, hora)
            return tiempo

    @staticmethod
    def __obtener_ruta_archivo(select: Literal['temp', 'registro'] = None) -> Union[str, Tuple[str, str]]:
        archivo_temp = ruta_carpeta_temp + '/Alumnos_' + ExcelManager.__tiempo().fecha + \
            '.xlsx'
        archivo_registro = ruta_carpeta_registro + \
            '/Alumnos_' + ExcelManager.__tiempo().fecha + '.xlsx'

        if select == 'temp':
            return archivo_temp
        elif select == 'registro':
            return archivo_registro
        else:
            RutaArchivo = namedtuple('RutaArchivo', ['temp', 'registro'])
            ruta_archivo = RutaArchivo(archivo_temp, archivo_registro)
            return ruta_archivo

    @staticmethod
    def eliminar_archivos_antiguos():
        nombre_base = "Alumnos_"
        extension = ".xlsx"
        diferencia_dias = 2

        fecha_actual = datetime.now().date()
        # Directorio ".temp" dentro de la carpeta actual
        directorio = path.join(getcwd(), ".temp")

        for archivo in listdir(directorio):
            if archivo.startswith(nombre_base) and archivo.endswith(extension):
                fecha_archivo_str = archivo.replace(
                    nombre_base, "").replace(extension, "")
                try:
                    fecha_archivo = datetime.strptime(
                        fecha_archivo_str, "%d_%m_%Y").date()
                    if (fecha_actual - fecha_archivo) > timedelta(days=diferencia_dias):
                        ruta_archivo = path.join(directorio, archivo)
                        remove(ruta_archivo)
                        dbg(f"Se ha eliminado el archivo: {archivo}")
                except ValueError:
                    # Ignorar archivos con nombres que no siguen el formato esperado
                    continue


class SerialManager:
    def __init__(self) -> None:
        dbg('Dentro...')
        excel_manager = ExcelManager()
        excel_manager.mkdirs()

        excel_manager.eliminar_archivos_antiguos()

        self.arduino: Serial
        self.datos_obtenidos: str = ''

        try:
            self.establecer_conexion()
        except KeyboardInterrupt:
            pass
        else:
            while True:
                try:
                    self.lectura()
                    print(self.datos_obtenidos)

                    excel_manager.save(self.datos_obtenidos)
                    excel_manager.copiar_hacia()

                except SerialException:
                    self.establecer_conexion()
                    self.reset_vars()
                    continue

                except KeyboardInterrupt:
                    self.cerrar_puerto()
                    break

                self.reset_vars()
        dbg('Saliendo...')

    def lectura(self) -> None:
        while self.datos_obtenidos.count('@') < 4:
            if self.arduino.in_waiting > 0:
                cadena_temp = self.arduino.readline()
                cadena_temp = str(cadena_temp, 'utf-8')
                cadena = ''
                for caracter in cadena_temp.split(' ')[:-1]:
                    cadena += chr(int(caracter, 16))
                self.datos_obtenidos += cadena
        self.datos_obtenidos = self.datos_obtenidos.replace('@', '\n')

    def establecer_conexion(self) -> None:
        i: int = 0
        try:
            while True:
                puerto = self.obtener_puerto()
                dbg(f'Intento No.{i + 1}')
                try:
                    self.arduino = Serial(puerto, 115200)

                except SerialException:
                    dbg('ERROR')
                    sleep(3)

                else:
                    dbg(f'Conexion establecida con: {puerto}')
                    return True

                i += 1
        except KeyboardInterrupt:
            dbg('intento de establecer conexion cancelado')
            raise KeyboardInterrupt()

    def cerrar_puerto(self) -> None:
        try:
            self.arduino.close()
        except SerialException:
            pass
        except AttributeError:
            pass
        print('Puerto cerrado')

    def reset_vars(self) -> None:
        self.datos_obtenidos: str = ''

    @staticmethod
    def obtener_puerto() -> str:
        for puerto in puertos.comports():
            if puerto.manufacturer in desarrolladores:
                return puerto[0]
        return ''


dbg('Iniciado')
SerialManager()
dbg('Finalizado')
